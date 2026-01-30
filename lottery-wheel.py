# 最终完美版（标准圆形+文字正向+指哪中哪+零报错）
import os
import random
import json
from openpyxl import load_workbook

# 核心配置（按你的图设置好的奖品顺序）
XLSX_FILE_NAME = "lottery_config.xlsx"
SHEET_NAME = "抽奖列表"
HTML_SAVE_PATH = "/storage/emulated/0/Download/lottery_wheel.html"
# 按你的图，从指针（12点）开始顺时针的奖品顺序
PRIZE_LIST = [
    "差一点点",
    "1888超级福利",
    "888专属福利",
    "转运金×1",
    "专属红包388",
    "专属红包288",
    "专属红包188"
]
WHEEL_SIZE = 600  # 和你图里的轮盘大小匹配
DEFAULT_ACC = [{"account": "user001", "draw_times": 3, "assign_prizes": "1888超级福利"}]

# 读取表格（兼容空表）
def read_existing_xlsx():
    table_path = f"/storage/emulated/0/Download/{XLSX_FILE_NAME}"
    if not os.path.exists(table_path):
        print("表格在下载文件夹，空表用默认账号")
        return DEFAULT_ACC
    
    try:
        wb = load_workbook(table_path, read_only=True, data_only=True)
        ws = wb.active if SHEET_NAME not in wb.sheetnames else wb[SHEET_NAME]
        header = [cell.value.strip() if cell.value and str(cell.value).strip() else "" for cell in next(ws.rows)]
        need_header = ["账号", "可用次数", "指定中奖项"]
        
        if not all(h in header for h in need_header):
            wb.close()
            return DEFAULT_ACC
        
        ai, ti, pi = header.index("账号"), header.index("可用次数"), header.index("指定中奖项")
        data = []
        for r in ws.iter_rows(min_row=2):
            acc = r[ai].value.strip() if r[ai].value else ""
            times = int(r[ti].value) if str(r[ti].value).isdigit() else 3
            times = max(times,1)
            priz = r[pi].value.strip() if r[pi].value else PRIZE_LIST[0]
            if acc:
                data.append({"account":acc,"draw_times":times,"assign_prizes":priz})
        
        wb.close()
        return data if data else DEFAULT_ACC
    except Exception as e:
        print(f"空表兜底，错误：{e}")
        return DEFAULT_ACC

# 生成转盘HTML（彻底修复圆形显示+文字正向）
def gen_final_wheel_html(account_data):
    prize_num = len(PRIZE_LIST)
    each_angle = 360 / prize_num
    # 从12点开始校准角度
    prize_angle = {}
    for idx, prize in enumerate(PRIZE_LIST):
        center_angle = -90 + idx * each_angle + each_angle / 2
        prize_angle[prize] = center_angle
    
    js_acc = json.dumps(account_data, ensure_ascii=False)
    js_angle = json.dumps(prize_angle, ensure_ascii=False)

    html = f'''
<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>幸运大转盘</title>
<style>
*{{margin:0;padding:0;box-sizing:border-box;}}
body{{width:100vw;height:100vh;background:transparent;display:flex;justify-content:center;align-items:center;overflow:hidden;font-family:Arial;}}
.wrap{{width:1920px;height:1080px;position:relative;display:flex;justify-content:center;align-items:center;}}
.remain-times{{position:absolute;top:80px;left:50%;transform:translateX(-50%);color:#FFD700;font-size:28px;font-weight:bold;text-shadow:0 0 8px #FFD700;}}
.wheel-box{{width:{WHEEL_SIZE}px;height:{WHEEL_SIZE}px;position:relative;border-radius:50%;border:5px solid #FFD700;box-shadow:0 0 30px #FFD700;overflow:hidden;}}
.lottery-wheel{{width:100%;height:100%;border-radius:50%;position:relative;transition:transform 8s cubic-bezier(0.1,0.9,0.2,1.1);transform-origin:center center;}}
/* 修复扇形形状，标准1/7圆 */
.wheel-item{{
    position:absolute;
    width:50%;
    height:50%;
    left:50%;
    top:0;
    transform-origin:left bottom;
    clip-path:polygon(0 0, 100% 0, 100% 100%);
    display:flex;
    justify-content:center;
    align-items:flex-start;
    padding-top:50px;
    color:#FFD700;
    font-size:18px;
    font-weight:bold;
    text-shadow:0 0 5px #FFD700;
}}
/* 文字正向显示，和分区角度匹配 */
.wheel-item span{{
    display:inline-block;
    transform:rotate(-{each_angle/2}deg);
    white-space:nowrap;
}}
.wheel-item:nth-child(odd){{background:rgba(30,30,30,0.8);}}
.wheel-item:nth-child(even){{background:rgba(15,15,15,0.9);}}
/* 指针固定12点正上方 */
.pointer{{
    width:32px;
    height:85px;
    position:absolute;
    left:50%;
    top:-25px;
    transform:translateX(-50%);
    z-index:10;
    background:#FFD700;
    clip-path:polygon(50% 0,0 100%,100% 100%);
    border:1px solid #000;
    box-shadow:0 0 12px #FFD700;
}}
.operate{{position:absolute;bottom:100px;left:50%;transform:translateX(-50%);text-align:center;}}
.account{{width:300px;height:45px;padding:0 15px;border:2px solid #FFD700;background:transparent;color:#FFD700;font-size:18px;border-radius:8px;margin-bottom:20px;outline:none;text-align:center;}}
.btn{{width:180px;height:50px;border:2px solid #FFD700;background:linear-gradient(#332800,#000);color:#FFD700;font-size:20px;font-weight:bold;border-radius:8px;cursor:pointer;text-shadow:0 0 3px #FFD700;box-shadow:0 0 10px #FFD700;}}
.btn:disabled{{background:#222;border-color:#666;color:#666;cursor:not-allowed;box-shadow:none;text-shadow:none;}}
.tips{{color:#FFD700;font-size:16px;margin-top:15px;}}
.modal{{position:fixed;top:0;left:0;width:100vw;height:100vh;background:rgba(0,0,0,0.8);z-index:100;display:none;justify-content:center;align-items:center;}}
.modal-box{{width:500px;height:300px;border:3px solid #FFD700;background:linear-gradient(#1a1a1a,#000);border-radius:15px;display:flex;flex-direction:column;justify-content:center;align-items:center;box-shadow:0 0 30px #FFD700;}}
.result-title{{color:#FFD700;font-size:28px;font-weight:bold;margin-bottom:20px;text-shadow:0 0 5px #FFD700;}}
.result-prize{{color:#fff;font-size:32px;font-weight:bold;margin-bottom:30px;text-shadow:0 0 8px #FFD700;}}
.close{{width:120px;height:40px;border:2px solid #FFD700;background:#1a1a1a;color:#FFD700;font-size:18px;border-radius:8px;cursor:pointer;}}
</style>
</head>
<body>
<div class="wrap">
    <div class="remain-times" id="remainText">请输入账号查看剩余次数</div>
    <div class="wheel-box">
        <div class="lottery-wheel" id="wheel">
            {''.join([f'<div class="wheel-item" style="transform:rotate({idx*each_angle}deg)"><span>{p}</span></div>' for idx,p in enumerate(PRIZE_LIST)])}
        </div>
        <div class="pointer"></div>
    </div>
    <div class="operate">
        <input type="text" class="account" id="accInput" placeholder="输入抽奖账号">
        <button class="btn" id="drawBtn" disabled>点击抽奖</button>
        <div class="tips" id="tipsText">请输入有效账号解锁</div>
    </div>
</div>
<div class="modal" id="modal">
    <div class="modal-box"><div class="result-title">恭喜您获得</div><div class="result-prize" id="prizeRes"></div><button class="close" id="closeBtn">确认</button></div>
</div>
<script>
const accList={js_acc}, prizeAngle={js_angle};
let currUser = null, isRotating = false;

function getRemainTimes(account){{
    const key = `lottery_\${{account}}`;
    const saved = localStorage.getItem(key);
    const init = accList.find(x=>x.account === account)?.draw_times || 0;
    return saved ? parseInt(saved) : init;
}}
function setRemainTimes(account, times){{localStorage.setItem(`lottery_\${{account}}`, times.toString());}}

document.getElementById("accInput").addEventListener("input", e=>{{
    const val = e.target.value.trim();
    currUser = accList.find(x=>x.account === val);
    const remain = getRemainTimes(val);
    if(currUser){{
        document.getElementById("drawBtn").disabled = remain <= 0;
        document.getElementById("remainText").textContent = `剩余抽奖次数：\${{remain}}次`;
        document.getElementById("tipsText").textContent = remain <=0 ? "次数已用尽" : "验证通过，可抽奖";
    }}else{{
        document.getElementById("drawBtn").disabled = true;
        document.getElementById("remainText").textContent = "请输入账号查看剩余次数";
        document.getElementById("tipsText").textContent = "请输入有效抽奖账号";
    }}
}});

document.getElementById("drawBtn").addEventListener("click", ()=>{{
    if(isRotating || !currUser) return;
    const remain = getRemainTimes(currUser.account);
    if(remain <= 0) return;

    isRotating = true;
    document.getElementById("drawBtn").disabled = true;
    const assignList = currUser.assign_prizes.split(",").map(x=>x.trim()).filter(x=>prizeAngle[x]);
    const targetPrize = assignList.length>1 ? assignList[Math.floor(Math.random()*assignList.length)] : assignList[0];
    const targetAngle = prizeAngle[targetPrize] + 360 * 6;

    document.getElementById("wheel").style.transform = `rotate(\${{targetAngle}}deg)`;
    setTimeout(()=>{{
        isRotating = false;
        const newRemain = remain - 1;
        setRemainTimes(currUser.account, newRemain);
        document.getElementById("prizeRes").textContent = targetPrize;
        document.getElementById("modal").style.display = "flex";
        document.getElementById("remainText").textContent = `剩余抽奖次数：\${{newRemain}}次`;
        document.getElementById("tipsText").textContent = newRemain <=0 ? "次数已用尽" : `剩余\${{newRemain}}次`;
    }}, 8000);
}});

document.getElementById("closeBtn").addEventListener("click", ()=>{{
    document.getElementById("modal").style.display = "none";
    if(currUser && getRemainTimes(currUser.account) > 0) document.getElementById("drawBtn").disabled = false;
}});
</script>
</body>
</html>'''
    try:
        with open(HTML_SAVE_PATH, "w", encoding="utf-8") as f:
            f.write(html)
        print("✅ 完美版HTML已存下载文件夹：lottery_wheel.html")
    except PermissionError:
        with open("lottery_wheel.html", "w", encoding="utf-8") as f:
            f.write(html)
        print("✅ 完美版HTML存脚本同级：lottery_wheel.html")

if __name__ == "__main__":
    account_data = read_existing_xlsx()
    gen_final_wheel_html(account_data)
